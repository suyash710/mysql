import xlsxwriter
from xlutils.copy import copy
from xlrd import open_workbook
import os.path
from openpyxl import load_workbook
import openpyxl
import os

from sys import argv as s     #reading the version input from user
string1='./DDL/'; string2=s[1]; string=string1+string2            #path to the version 
os.chdir(string)
files=string;
for i in os.listdir(os.getcwd()):
        if i.endswith(".sql"):
             print("\n\n",i);

             with open(i) as f:
                  sqlContent = f.read().splitlines()

                  subType= "#Type"; 
                  for typeData in sqlContent:       #for loop to find the "Type" of SQL data file. 
                        
                        if subType in typeData:
                              typeData= typeData.split();
                              Type=typeData[len(typeData)-1];           
                              print("TYPE-----",Type)
                              if (Type.lower()== "master" or Type.lower()== "constant"):  
                                    Type=Type.lower();
                              else:
                                    Type="transactional";
                              print("hello---------");
                              print("\nType is - ",Type,"\n");
                        
                        # else:
                        #       Type="transactional";
                        #       print("\nType is - ",Type,"\n");


                  subName="CREATE";
                  for nameData in sqlContent:               #for loop to get the SQL file Name.
                        if subName in nameData:
                              nameData=nameData.split();
                              Name=nameData[len(nameData)-2]; name_part=Name.split(".");
                              Name=name_part[len(name_part)-1]; name_part=Name.split("`");
                              Name=name_part[len(name_part)-2];
                              print("\nFile_name - ",Name,"\n");        #prints the SQL file Name
                        
                        # else:
                        #       break;

                  if os.path.isfile("../../dataSQL.xlsx"):
                        
                        print ("file exist");
                        wb = load_workbook("../../dataSQL.xlsx", use_iterators=True)
                        sheet = wb.worksheets[0]
                        row_count = sheet.get_highest_row()
                        print ("row count: ",row_count);

                        wb = openpyxl.load_workbook("../../dataSQL.xlsx")
                        ws = wb.get_sheet_by_name('Sheet1')
                        ws.cell(row=row_count+1, column=1, value=row_count)
                        ws.cell(row=row_count+1, column=2, value=Name)
                        ws.cell(row=row_count+1, column=3, value=Type)
                        wb.save("../../dataSQL.xlsx")

                  else:
                        print ("file dosent exist, making a new Sheet");
                        workbook=xlsxwriter.Workbook("../../dataSQL.xlsx");   # makes an Excel Workbook     
                        print (os.getcwd())
                        worksheet=workbook.add_worksheet();       # adds a Sheet to the Workbook

                        worksheet.write(0,0,"#");                 # adds the Index to the excel sheet.
                        worksheet.write(0,2,"Type");
                        worksheet.write(0,1,"Name");

                        worksheet.write(1,2,Type);                #writes the Type to sheet.          
                        worksheet.write(1,1,Name);                #writes the Name to sheet.
                        worksheet.write(1,0,1);

                        workbook.close();       #Name of the Excel Sheet